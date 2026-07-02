import io
from datetime import date as dt_date
from itertools import combinations

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

FILL_GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFDAA5", end_color="FFDAA5", fill_type="solid")
FILL_RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_LGRAY  = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
FILL_HEADER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_BLUE   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
FILL_SECTION= PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

FONT_BOLD   = Font(bold=True)
FONT_HEADER = Font(bold=True, size=11)
FONT_SMALL  = Font(size=9)

MAX_GROUP_SIZE = 5
MAX_CANDIDATES = 20

EXCEL_STRINGS = {
    "ru": {
        "sheet":      "Сверка",
        "title":      "ИТОГ СВЕРКИ",
        "total_a":    "Итого сумма А:",
        "total_b":    "Итого сумма Б:",
        "diff":       "Разница итогов:",
        "rows_a":     "Строк в А:",
        "rows_b":     "Строк в Б:",
        "pairs":      "Совпавших пар:",
        "unm_a_lbl":  "Расхождений А:",
        "unm_b_lbl":  "Расхождений Б:",
        "legend":     "Легенда:",
        "leg_exact":  "Точное совпадение",
        "leg_tol":    "По дате{}",
        "leg_group":  "Групповое совпадение",
        "leg_disc":   "Расхождение",
        "col_date_a": "Дата А",
        "col_amt_a":  "Сумма А",
        "col_date_b": "Дата Б",
        "col_amt_b":  "Сумма Б",
        "col_status": "Статус",
        "sec_match":  "▼ СОВПАДЕНИЯ",
        "sec_disc":   "▼ РАСХОЖДЕНИЯ",
        "tol_wday":   " (раб. день)",
        "tol_days":   " (±{} дн.)",
        "st_exact":   "Совпадение",
        "st_tol":     "Совпадение (дата ±{} дн.)",
        "st_group_a": "Группа А→Б",
        "st_group_b": "Группа Б→А",
        "st_unm_a":   "Только в А",
        "st_unm_b":   "Только в Б",
    },
    "en": {
        "sheet":      "Reconciliation",
        "title":      "RECONCILIATION SUMMARY",
        "total_a":    "Total amount A:",
        "total_b":    "Total amount B:",
        "diff":       "Total difference:",
        "rows_a":     "Rows in A:",
        "rows_b":     "Rows in B:",
        "pairs":      "Matched pairs:",
        "unm_a_lbl":  "Discrepancies A:",
        "unm_b_lbl":  "Discrepancies B:",
        "legend":     "Legend:",
        "leg_exact":  "Exact match",
        "leg_tol":    "By date{}",
        "leg_group":  "Group match",
        "leg_disc":   "Discrepancy",
        "col_date_a": "Date A",
        "col_amt_a":  "Amount A",
        "col_date_b": "Date B",
        "col_amt_b":  "Amount B",
        "col_status": "Status",
        "sec_match":  "▼ MATCHES",
        "sec_disc":   "▼ DISCREPANCIES",
        "tol_wday":   " (workday)",
        "tol_days":   " (±{} d.)",
        "st_exact":   "Match",
        "st_tol":     "Match (date ±{} d.)",
        "st_group_a": "Group A→B",
        "st_group_b": "Group B→A",
        "st_unm_a":   "Only in A",
        "st_unm_b":   "Only in B",
    },
}


RU_MONTHS = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "мая": 5,
    "июн": 6, "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
}


def _parse_dates(series: pd.Series) -> pd.Series:
    """Parse dates supporting:
    - Proper datetime (from Excel date cells) → used as-is
    - Float like 1.07 or 30.06 → treated as day.month, year = current
    - String like '1.07' (one dot) → same d.mm interpretation
    - String like '3 мар' or '03 марта 2025' → Russian month name
    - Full string '01.07.2025' (two dots) → standard parsing
    """
    cur_year = dt_date.today().year

    def parse_one(val):
        if pd.isna(val):
            return pd.NaT
        if isinstance(val, pd.Timestamp):
            return val
        # Numeric: interpret as d.mm float (e.g. 1.07 → July 1)
        if isinstance(val, (int, float)):
            try:
                return pd.to_datetime(f"{float(val):.2f}.{cur_year}", format="%d.%m.%Y")
            except Exception:
                return pd.NaT
        # String
        s = str(val).strip()
        # Russian month name: "3 мар", "03 марта", "3 марта 2025"
        parts = s.split()
        if len(parts) >= 2 and parts[0].isdigit():
            key = parts[1].lower()[:3]
            if key in RU_MONTHS:
                try:
                    day   = int(parts[0])
                    month = RU_MONTHS[key]
                    year  = int(parts[2]) if len(parts) >= 3 and parts[2].isdigit() else cur_year
                    return pd.Timestamp(year, month, day)
                except Exception:
                    pass
        if s.count(".") == 1:
            # Short format "d.mm" without year
            try:
                return pd.to_datetime(f"{s}.{cur_year}", format="%d.%m.%Y")
            except Exception:
                pass
        # Full date string — standard parsing
        try:
            return pd.to_datetime(s, dayfirst=True)
        except Exception:
            return pd.NaT

    return series.apply(parse_one)


def load_df(file_bytes: bytes, date_col: str, amount_col: str, has_header: bool = True) -> pd.DataFrame:
    if has_header:
        df = pd.read_excel(io.BytesIO(file_bytes))
        df.columns = [str(c) for c in df.columns]
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), header=None)
        df.columns = [f"Колонка {i + 1}" for i in range(len(df.columns))]

    missing = [c for c in (date_col, amount_col) if c not in df.columns]
    if missing:
        raise ValueError(f"Колонки не найдены: {', '.join(missing)}")

    df = df[[date_col, amount_col]].copy()
    df.columns = ["date", "amount"]
    raw_date_samples = [str(x) for x in df["date"].head(5).tolist()]
    df["date"] = _parse_dates(df["date"])

    # Robust amount parsing: handle spaces/nbsp as thousands separator, comma as decimal
    raw = df["amount"].astype(str).str.replace(r"[\s ]", "", regex=True).str.replace(",", ".", regex=False)
    df["amount"] = pd.to_numeric(raw, errors="coerce")

    bad_dates      = int(df["date"].isna().sum())
    bad_amounts    = int(df["amount"].isna().sum())
    date_samples   = df["date"].dropna().head(3).dt.strftime("%d.%m.%Y").tolist()
    amount_samples = [round(x, 2) for x in df["amount"].dropna().head(3).tolist()]

    df = df.dropna(subset=["date", "amount"]).reset_index(drop=True)

    if df.empty:
        raise ValueError(
            f"Нет данных после парсинга. "
            f"Невалидных дат: {bad_dates}, невалидных сумм: {bad_amounts}. "
            f"Примеры распознанных дат: {date_samples or 'нет'}. "
            f"Исходные значения дат: {raw_date_samples}. "
            f"Примеры распознанных сумм: {amount_samples or 'нет'}."
        )

    return df


def _workday_distance(d1, d2) -> int:
    """Calendar days between d1 and d2, but Friday→Monday counts as 1."""
    delta = abs((d1 - d2).days)
    if delta == 3:
        later, earlier = (d1, d2) if d1 > d2 else (d2, d1)
        if later.weekday() == 0 and earlier.weekday() == 4:  # Mon=0, Fri=4
            return 1
    return delta


def _within(d1, d2, tol: int) -> bool:
    # tol == -1 means "1 workday" mode
    if tol == -1:
        return _workday_distance(d1, d2) <= 1
    return abs((d1 - d2).days) <= tol


def _match(df_a: pd.DataFrame, df_b: pd.DataFrame, tol: int):
    """4-phase matching.
    Order: exact → group A→B → group B→A → tolerance 1:1
    Groups are found before tolerance matches to avoid 'stealing' rows
    that should participate in a group.
    """
    n_a, n_b = len(df_a), len(df_b)
    matched_a: set = set()
    matched_b: set = set()
    pairs: list = []

    # ── Phase 1: exact date + exact amount ──────────────────────
    for i in range(n_a):
        for j in range(n_b):
            if j in matched_b:
                continue
            if (df_a.at[i, "date"] == df_b.at[j, "date"] and
                    abs(df_a.at[i, "amount"] - df_b.at[j, "amount"]) < 0.005):
                pairs.append({"a": [i], "b": [j], "type": "exact"})
                matched_a.add(i)
                matched_b.add(j)
                break

    # ── Phase 2: group A→B (one A row = sum of several B rows) ──
    for i in range(n_a):
        if i in matched_a:
            continue
        a_amt  = df_a.at[i, "amount"]
        a_date = df_a.at[i, "date"]
        cands  = [
            j for j in range(n_b)
            if j not in matched_b and _within(df_b.at[j, "date"], a_date, tol)
        ][:MAX_CANDIDATES]
        found = False
        for size in range(2, min(MAX_GROUP_SIZE + 1, len(cands) + 1)):
            for combo in combinations(cands, size):
                if abs(sum(df_b.at[j, "amount"] for j in combo) - a_amt) < 0.005:
                    pairs.append({"a": [i], "b": list(combo), "type": "group_a"})
                    matched_a.add(i)
                    matched_b.update(combo)
                    found = True
                    break
            if found:
                break

    # ── Phase 3: group B→A (one B row = sum of several A rows) ──
    for j in range(n_b):
        if j in matched_b:
            continue
        b_amt  = df_b.at[j, "amount"]
        b_date = df_b.at[j, "date"]
        cands  = [
            i for i in range(n_a)
            if i not in matched_a and _within(df_a.at[i, "date"], b_date, tol)
        ][:MAX_CANDIDATES]
        found = False
        for size in range(2, min(MAX_GROUP_SIZE + 1, len(cands) + 1)):
            for combo in combinations(cands, size):
                if abs(sum(df_a.at[i, "amount"] for i in combo) - b_amt) < 0.005:
                    pairs.append({"a": list(combo), "b": [j], "type": "group_b"})
                    matched_b.add(j)
                    matched_a.update(combo)
                    found = True
                    break
            if found:
                break

    # ── Phase 4: tolerance 1:1 (same amount, date within ±tol) ──
    if tol != 0:
        for i in range(n_a):
            if i in matched_a:
                continue
            for j in range(n_b):
                if j in matched_b:
                    continue
                if (_within(df_a.at[i, "date"], df_b.at[j, "date"], tol) and
                        abs(df_a.at[i, "amount"] - df_b.at[j, "amount"]) < 0.005):
                    pairs.append({"a": [i], "b": [j], "type": "tolerance"})
                    matched_a.add(i)
                    matched_b.add(j)
                    break

    return pairs, matched_a, matched_b


def _pair_min_date(p, df_a, df_b):
    dates = [df_a.at[i, "date"] for i in p["a"]] + [df_b.at[j, "date"] for j in p["b"]]
    return min(dates)


def reconcile_files(
    bytes_a: bytes, bytes_b: bytes,
    date_col_a: str, amount_col_a: str,
    date_col_b: str, amount_col_b: str,
    date_tolerance: str,
    has_header_a: bool = True,
    has_header_b: bool = True,
    lang: str = "ru",
) -> tuple[bytes, dict]:
    workday = (str(date_tolerance).strip() == "workday")
    tol = -1 if workday else int(date_tolerance)

    df_a = load_df(bytes_a, date_col_a, amount_col_a, has_header_a)
    df_b = load_df(bytes_b, date_col_b, amount_col_b, has_header_b)

    if df_a.empty:
        raise ValueError("Файл А не содержит данных после парсинга")
    if df_b.empty:
        raise ValueError("Файл Б не содержит данных после парсинга")

    pairs, matched_a, matched_b = _match(df_a, df_b, tol)

    unm_a = sorted(
        [i for i in range(len(df_a)) if i not in matched_a],
        key=lambda i: df_a.at[i, "date"],
    )
    unm_b = sorted(
        [j for j in range(len(df_b)) if j not in matched_b],
        key=lambda j: df_b.at[j, "date"],
    )

    total_a = float(df_a["amount"].sum())
    total_b = float(df_b["amount"].sum())
    summary = {
        "total_a":     total_a,
        "total_b":     total_b,
        "diff":        round(total_a - total_b, 2),
        "rows_a":      len(df_a),
        "rows_b":      len(df_b),
        "total_pairs": len(pairs),
        "exact":       sum(1 for p in pairs if p["type"] == "exact"),
        "tolerance":   sum(1 for p in pairs if p["type"] == "tolerance"),
        "group":       sum(1 for p in pairs if p["type"] in ("group_a", "group_b")),
        "unmatched_a": len(unm_a),
        "unmatched_b": len(unm_b),
        "workday":     workday,
    }

    s  = EXCEL_STRINGS.get(lang, EXCEL_STRINGS["ru"])
    wb = _build_excel(df_a, df_b, pairs, unm_a, unm_b, summary, tol, s)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), summary


def _build_excel(df_a, df_b, pairs, unm_a, unm_b, summary, tol, s: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = s["sheet"]

    tol_label = s["tol_wday"] if tol == -1 else (s["tol_days"].format(tol) if tol else "")

    def sc(row, col, val, **kw):
        c = ws.cell(row, col, val)
        for attr, v in kw.items():
            setattr(c, attr, v)
        return c

    # ── Summary ──────────────────────────────────────────────────
    ws.append([s["title"]])
    r = ws.max_row
    sc(r, 1, s["title"], font=Font(bold=True, size=13), fill=FILL_BLUE)
    for col in range(2, 6):
        ws.cell(r, col).fill = FILL_BLUE

    fmt = "#,##0.00"
    ws.append([s["total_a"], summary["total_a"], "", s["total_b"], summary["total_b"]])
    r = ws.max_row
    sc(r, 1, s["total_a"], font=FONT_BOLD); ws.cell(r, 2).number_format = fmt
    sc(r, 4, s["total_b"], font=FONT_BOLD); ws.cell(r, 5).number_format = fmt

    diff = summary["diff"]
    ws.append([s["diff"], diff])
    r = ws.max_row
    sc(r, 1, s["diff"], font=FONT_BOLD); ws.cell(r, 2).number_format = fmt
    if abs(diff) > 0.005:
        ws.cell(r, 2).fill = FILL_RED

    ws.append([s["rows_a"], summary["rows_a"], "", s["rows_b"], summary["rows_b"]])
    r = ws.max_row
    sc(r, 1, s["rows_a"], font=FONT_BOLD); sc(r, 4, s["rows_b"], font=FONT_BOLD)

    ws.append([
        s["pairs"], summary["total_pairs"], "",
        s["unm_a_lbl"], summary["unmatched_a"],
        s["unm_b_lbl"], summary["unmatched_b"],
    ])
    r = ws.max_row
    sc(r, 1, s["pairs"],     font=FONT_BOLD)
    sc(r, 4, s["unm_a_lbl"], font=FONT_BOLD)
    sc(r, 6, s["unm_b_lbl"], font=FONT_BOLD)
    if summary["unmatched_a"]: ws.cell(r, 5).fill = FILL_RED
    if summary["unmatched_b"]: ws.cell(r, 7).fill = FILL_RED

    ws.append([])

    # ── Legend ───────────────────────────────────────────────────
    ws.append([s["legend"]])
    sc(ws.max_row, 1, s["legend"], font=FONT_BOLD)
    lr = ws.max_row + 1
    for col, (fill, label) in enumerate([
        (FILL_GREEN,  s["leg_exact"]),
        (FILL_YELLOW, s["leg_tol"].format(tol_label)),
        (FILL_ORANGE, s["leg_group"]),
        (FILL_RED,    s["leg_disc"]),
    ], 1):
        c = ws.cell(lr, col, label)
        c.fill = fill; c.font = FONT_SMALL
    ws.append([])

    # ── Data header ──────────────────────────────────────────────
    ws.append([s["col_date_a"], s["col_amt_a"], s["col_date_b"], s["col_amt_b"], s["col_status"]])
    hdr_row = ws.max_row
    for col in range(1, 6):
        c = ws.cell(hdr_row, col)
        c.font = FONT_HEADER; c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:E{hdr_row}"

    def write_row(date_a, amt_a, date_b, amt_b, status, fill, secondary=False):
        ws.append([
            date_a if date_a is not None else "",
            amt_a  if amt_a  is not None else "",
            date_b if date_b is not None else "",
            amt_b  if amt_b  is not None else "",
            status,
        ])
        r = ws.max_row
        row_fill = FILL_LGRAY if secondary else fill
        for col in range(1, 6):
            ws.cell(r, col).fill = row_fill
        if isinstance(amt_a, (int, float)):
            ws.cell(r, 2).number_format = fmt
        if isinstance(amt_b, (int, float)):
            ws.cell(r, 4).number_format = fmt

    def section_header(label, fill):
        ws.append([label])
        r = ws.max_row
        sc(r, 1, label, font=FONT_BOLD, fill=fill)
        for col in range(2, 6):
            ws.cell(r, col).fill = fill

    FILLS = {
        "exact": FILL_GREEN, "tolerance": FILL_YELLOW,
        "group_a": FILL_ORANGE, "group_b": FILL_ORANGE,
    }

    # ── Matched pairs ────────────────────────────────────────────
    section_header(s["sec_match"], FILL_SECTION)
    for p in sorted(pairs, key=lambda p: _pair_min_date(p, df_a, df_b)):
        a_idxs, b_idxs, ptype = p["a"], p["b"], p["type"]
        fill = FILLS[ptype]
        if ptype == "tolerance":
            label = s["st_tol"].format(tol)
        elif ptype == "group_a":
            label = s["st_group_a"]
        elif ptype == "group_b":
            label = s["st_group_b"]
        else:
            label = s["st_exact"]
        for k in range(max(len(a_idxs), len(b_idxs))):
            i = a_idxs[k] if k < len(a_idxs) else None
            j = b_idxs[k] if k < len(b_idxs) else None
            write_row(
                df_a.at[i, "date"].strftime("%d.%m.%Y") if i is not None else None,
                df_a.at[i, "amount"] if i is not None else None,
                df_b.at[j, "date"].strftime("%d.%m.%Y") if j is not None else None,
                df_b.at[j, "amount"] if j is not None else None,
                label if k == 0 else "",
                fill,
                secondary=(k > 0),
            )

    # ── Unmatched (interleaved by date) ──────────────────────────
    if unm_a or unm_b:
        ws.append([])
        section_header(s["sec_disc"], FILL_RED)
        unmatched_all = (
            [("a", i, df_a.at[i, "date"]) for i in unm_a] +
            [("b", j, df_b.at[j, "date"]) for j in unm_b]
        )
        for which, idx, _ in sorted(unmatched_all, key=lambda x: x[2]):
            if which == "a":
                write_row(
                    df_a.at[idx, "date"].strftime("%d.%m.%Y"), df_a.at[idx, "amount"],
                    None, None, s["st_unm_a"], FILL_RED,
                )
            else:
                write_row(
                    None, None,
                    df_b.at[idx, "date"].strftime("%d.%m.%Y"), df_b.at[idx, "amount"],
                    s["st_unm_b"], FILL_RED,
                )

    # ── Column widths ─────────────────────────────────────────────
    for col, width in zip("ABCDE", [13, 16, 13, 16, 30]):
        ws.column_dimensions[col].width = width

    return wb
